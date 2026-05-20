import subprocess
import json
import os
import configparser
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def run_oci_command(cmd_args, allow_empty=False):
    result = subprocess.run(cmd_args, capture_output=True, text=True)
    if result.returncode != 0:
        print("❌ OCI CLI command failed:")
        print("Command:", " ".join(cmd_args))
        print("stderr:", result.stderr.strip())
        exit(1)
    if not result.stdout.strip():
        if allow_empty:
            return None
        print("❌ OCI CLI returned no output.")
        print("Command:", " ".join(cmd_args))
        exit(1)
    try:
        return json.loads(result.stdout)
    except json.JSONDecodeError as e:
        print("❌ Failed to parse JSON.")
        print("Output was:", result.stdout.strip())
        print("Error:", e)
        exit(1)


def get_tenancy_ocid_from_config():
    config_path = os.path.expanduser("~/.oci/config")
    if not os.path.exists(config_path):
        print("❌ OCI config file not found at ~/.oci/config")
        exit(1)
    config = configparser.ConfigParser()
    config.read(config_path)
    if 'DEFAULT' not in config or 'tenancy' not in config['DEFAULT']:
        print("❌ Tenancy OCID not found in ~/.oci/config under DEFAULT profile")
        exit(1)
    return config['DEFAULT']['tenancy']


def fetch_compartments(tenancy_ocid):
    compartments = {}

    def _fetch(compartment_id):
        data = run_oci_command([
            "oci", "iam", "compartment", "list",
            "--compartment-id", compartment_id,
            "--all",
            "--query", 'data[?"lifecycle-state"==`ACTIVE`]'
        ], allow_empty=True)
        if not data:
            return
        for c in data:
            comp_id = c.get("id")
            comp_name = c.get("name", "Unknown")
            compartments[comp_id] = comp_name
            _fetch(comp_id)

    compartments[tenancy_ocid] = "Root Compartment"
    _fetch(tenancy_ocid)
    return compartments


def fetch_policies(compartment_ids):
    policies = []
    for comp_id in compartment_ids:
        print(f"🔍 Fetching policies in compartment: {comp_id}")
        data = run_oci_command([
            "oci", "iam", "policy", "list",
            "--compartment-id", comp_id,
            "--all",
            "--query", 'data[?"lifecycle-state"==`ACTIVE`]'
        ], allow_empty=True)
        if not data:
            print(f"  No policies found in compartment {comp_id}")
            continue
        for policy in data:
            policies.append({
                "policy_name": policy.get("name"),
                "policy_id": policy.get("id"),
                "compartment_id": comp_id,
                "policy_statements": "\n".join(policy.get("statements", []))
            })
    return policies


def write_policies_to_excel(policies, filename="oci_policies.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "OCI Policies"

    headers = ["Policy Name", "Policy ID", "Compartment OCID", "Policy Statements"]
    ws.append(headers)

    for pol in policies:
        ws.append([
            pol["policy_name"],
            pol["policy_id"],
            pol["compartment_id"],
            pol["policy_statements"]
        ])

    for i, col in enumerate(headers, 1):
        max_length = max(
            (len(str(row[i-1])) for row in ws.iter_rows(min_row=2, values_only=True)),
            default=0
        )
        max_length = max(max_length, len(col)) + 2
        ws.column_dimensions[get_column_letter(i)].width = max_length

    wb.save(filename)
    print(f"✅ Policies written to {filename}")


def main():
    print("🔍 Reading tenancy OCID from ~/.oci/config ...")
    tenancy_ocid = get_tenancy_ocid_from_config()
    print(f"✅ Tenancy OCID: {tenancy_ocid}")

    print("\n📂 Fetching all compartments (including root)...")
    compartments = fetch_compartments(tenancy_ocid)
    print(f"✅ Found {len(compartments)} compartments.")

    all_compartment_ids = list(compartments.keys())
    policies = fetch_policies(all_compartment_ids)

    if not policies:
        print("❌ No policies found for the compartments.")
        exit(0)

    write_policies_to_excel(policies)


if __name__ == "__main__":
    main()
